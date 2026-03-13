from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_required, current_user
from models import db, User, Vehicle, VehicleBooking, Driver, VehicleMileage, SystemConfig, DepartmentBudget
from sqlalchemy import and_, extract, or_
from datetime import datetime, date, timedelta
from views.telegram_service import (notify_approved, notify_forwarded_to_approver, notify_approver_approved, notify_rejected)
import os, time
from werkzeug.utils import secure_filename

vehicle_bp    = Blueprint('vehicle', __name__)
adminfleet_bp = Blueprint('adminfleet', __name__)
admincost_bp  = Blueprint('admincost', __name__)
driver_bp     = Blueprint('driver', __name__)


def is_vehicle_admin():
    return current_user.role_vehicle == 'admin' or current_user.is_superadmin


# ─────────────────────────────────────────────
# หน้าหลัก
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle')
@login_required
def index():
    bookings = VehicleBooking.query.order_by(VehicleBooking.created_at.desc()).all()
    return render_template('vehicle/vehicle.html', bookings=bookings)


# ─────────────────────────────────────────────
# จองรถแบบใหม่ — ไม่ต้องเลือกรถ admin กำหนดให้
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/book', methods=['POST'])
@login_required
def book_vehicle_simple():
    try:
        start_str = request.form.get('start_datetime', '').strip()
        end_str   = request.form.get('end_datetime',   '').strip()

        if not start_str or not end_str:
            flash('กรุณาเลือกวัน-เวลาไปและกลับให้ครบ', 'warning')
            return redirect(url_for('vehicle.index'))

        start_dt        = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
        end_dt          = datetime.strptime(end_str,   '%Y-%m-%dT%H:%M')
        destination     = request.form.get('destination')
        purpose         = request.form.get('purpose')
        passenger_count = int(request.form.get('passenger_count', 1))
        need_driver     = request.form.get('need_driver') == 'on'
        allow_join      = request.form.get('allow_join')  == 'on'

        if start_dt >= end_dt:
            flash('เวลากลับต้องมากกว่าเวลาไป', 'danger')
            return redirect(url_for('vehicle.index'))

        new_booking = VehicleBooking(
            user_id         = current_user.id,
            start_datetime  = start_dt,
            end_datetime    = end_dt,
            destination     = destination,
            purpose         = purpose,
            passenger_count = passenger_count,
            need_driver     = need_driver,
            allow_join      = allow_join,
            status          = 'pending'
        )
        db.session.add(new_booking)
        db.session.commit()
        flash(f'ส่งคำขอจองรถเรียบร้อยแล้ว (#{ new_booking.id }) รอ Admin พิจารณา', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return redirect(url_for('vehicle.index'))



# ─────────────────────────────────────────────
# แก้ไขการจอง
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/edit/<int:booking_id>', methods=['GET', 'POST'])
@login_required
def edit_booking(booking_id):
    booking = VehicleBooking.query.get_or_404(booking_id)

    if current_user.id != booking.user_id:
        flash('คุณไม่มีสิทธิ์แก้ไขรายการนี้', 'danger')
        return redirect(url_for('vehicle.index'))

    if request.method == 'POST':
        try:
            booking.start_datetime  = datetime.strptime(request.form.get('start_datetime'), '%Y-%m-%dT%H:%M')
            booking.end_datetime    = datetime.strptime(request.form.get('end_datetime'),   '%Y-%m-%dT%H:%M')
            booking.destination     = request.form.get('destination')
            booking.purpose         = request.form.get('purpose')
            booking.passenger_count = int(request.form.get('passenger_count', 1))
            booking.need_driver     = True if request.form.get('need_driver') else False
            booking.allow_join      = True if request.form.get('allow_join')  else False

            db.session.commit()
            flash('อัปเดตข้อมูลการจองเรียบร้อยแล้ว', 'success')
            return redirect(url_for('vehicle.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    start_str = booking.start_datetime.strftime('%Y-%m-%dT%H:%M')
    end_str   = booking.end_datetime.strftime('%Y-%m-%dT%H:%M')
    return render_template('vehicle/vehicle_edit.html', booking=booking, start_str=start_str, end_str=end_str)


# ─────────────────────────────────────────────
# ลบการจอง
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/delete/<int:booking_id>', methods=['POST'])
@login_required
def delete_booking(booking_id):
    booking = VehicleBooking.query.get_or_404(booking_id)

    if current_user.id != booking.user_id:
        flash('คุณไม่มีสิทธิ์ลบรายการนี้', 'danger')
        return redirect(url_for('vehicle.index'))

    try:
        db.session.delete(booking)
        db.session.commit()
        flash('ยกเลิกและลบรายการจองเรียบร้อยแล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'เกิดข้อผิดพลาดในการลบ: {str(e)}', 'danger')

    return redirect(url_for('vehicle.index'))


# ─────────────────────────────────────────────
# รายละเอียดทริป
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/detail/<int:booking_id>')
@login_required
def detail_booking(booking_id):
    booking = VehicleBooking.query.get_or_404(booking_id)

    if current_user.role_vehicle not in ['approver', 'admin'] and not current_user.is_superadmin and current_user.id != booking.user_id:
        flash('คุณไม่มีสิทธิ์เข้าถึงข้อมูลนี้', 'danger')
        return redirect(url_for('vehicle.index'))

    drivers = Driver.query.filter_by(is_active=True).all()
    return render_template('vehicle/vehicle_detail.html', booking=booking, drivers=drivers)


# ─────────────────────────────────────────────
# ปฏิทิน + API
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/calendar')
@login_required
def calendar_view():
    session['vehicle_back_url'] = url_for('vehicle.calendar_view')
    return render_template('vehicle/vehicle_calendar.html')


@vehicle_bp.route('/api/vehicle/bookings')
@login_required
def api_bookings():
    bookings = VehicleBooking.query.filter(
        VehicleBooking.status.in_(['pending', 'waiting_approver', 'approved'])
    ).all()

    events = []
    for b in bookings:
        color = '#198754' if b.status == 'approved' else ('#0dcaf0' if b.status == 'waiting_approver' else '#ffc107')
        vehicle_label = f"{b.assigned_vehicle.brand} {b.assigned_vehicle.model}" if b.assigned_vehicle else "รอกำหนดรถ"
        events.append({
            'id':    b.id,
            'title': f"{vehicle_label} ({b.user.full_name or b.user.username})",
            'start': b.start_datetime.isoformat() if b.start_datetime else None,
            'end':   b.end_datetime.isoformat()   if b.end_datetime   else None,
            'color': color,
            'url':   url_for('vehicle.detail_booking', booking_id=b.id)
        })
    return jsonify(events)


# ─────────────────────────────────────────────
# อนุมัติ / ปฏิเสธ
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/approve/<int:booking_id>', methods=['POST'])
@login_required
def approve_booking(booking_id):
    booking   = VehicleBooking.query.get_or_404(booking_id)
    action    = request.form.get('action')
    driver_id = request.form.get('driver_id')

    try:
        if is_vehicle_admin():
            if action == 'approve':
                booking.status = 'approved'
                if driver_id: booking.driver_id = driver_id
                db.session.commit()
                notify_approved(booking)
                flash('อนุมัติการจองรถเรียบร้อย', 'success')
            elif action == 'forward':
                booking.status = 'waiting_approver'
                if driver_id: booking.driver_id = driver_id
                db.session.commit()
                notify_forwarded_to_approver(booking)
                flash(f'ส่งต่อให้ Approver ของแผนก {booking.user.department} พิจารณาแล้ว', 'info')
            elif action == 'reject':
                booking.status = 'rejected'
                db.session.commit()
                notify_rejected(booking, current_user)
                flash('ไม่อนุมัติการจองรถ', 'danger')

        elif current_user.role_vehicle == 'approver':
            if booking.status != 'waiting_approver':
                flash('รายการนี้ไม่ได้อยู่ในสถานะรอคุณอนุมัติ', 'warning')
                return redirect(url_for('vehicle.detail_booking', booking_id=booking.id))

            if current_user.department != booking.user.department:
                flash('คุณสามารถอนุมัติได้เฉพาะแผนกเดียวกันเท่านั้น', 'danger')
                return redirect(url_for('vehicle.detail_booking', booking_id=booking.id))

            if action == 'approve':
                booking.status = 'approved'
                db.session.commit()
                notify_approver_approved(booking, current_user)
                flash('Approver อนุมัติการเดินทางเรียบร้อยแล้ว', 'success')
            elif action == 'reject':
                booking.status = 'rejected'
                db.session.commit()
                notify_rejected(booking, current_user)
                flash('Approver ปฏิเสธการเดินทางนี้', 'danger')
        else:
            flash('คุณไม่มีสิทธิ์ทำรายการนี้', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return redirect(url_for('vehicle.detail_booking', booking_id=booking.id))


# ─────────────────────────────────────────────
# Admin: จัดการรถและคนขับ
# ─────────────────────────────────────────────
@adminfleet_bp.route('/admin/manage-fleet', methods=['GET', 'POST'])
@login_required
def manage_fleet():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add_vehicle':
            new_vehicle = Vehicle(
                brand         = request.form.get('brand'),
                model         = request.form.get('model'),
                license_plate = request.form.get('license_plate'),
                capacity      = int(request.form.get('capacity')),
                fuel_rate     = float(request.form.get('fuel_rate') or 10)
            )
            db.session.add(new_vehicle)
            db.session.commit()
            flash(f"เพิ่มรถ {new_vehicle.brand} {new_vehicle.model} สำเร็จ!", 'success')

        elif action == 'add_driver':
            new_driver = Driver(
                name    = request.form.get('name'),
                phone   = request.form.get('phone'),
                user_id = request.form.get('user_id') or None
            )
            db.session.add(new_driver)
            db.session.commit()
            flash(f"เพิ่มพนักงานขับรถ {new_driver.name} สำเร็จ!", 'success')

        elif action == 'edit_vehicle':
            vid     = int(request.form.get('vehicle_id'))
            vehicle = Vehicle.query.get_or_404(vid)
            vehicle.brand         = request.form.get('brand')
            vehicle.model         = request.form.get('model')
            vehicle.license_plate = request.form.get('license_plate')
            vehicle.capacity      = int(request.form.get('capacity'))
            vehicle.status        = request.form.get('status', 'active')
            fuel_rate_str = request.form.get('fuel_rate', '').strip()
            if fuel_rate_str:
                vehicle.fuel_rate = float(fuel_rate_str)
            db.session.commit()
            flash(f"อัปเดตข้อมูลรถ {vehicle.brand} {vehicle.model} สำเร็จ!", 'success')

        elif action == 'delete_vehicle':
            vid     = int(request.form.get('vehicle_id'))
            vehicle = Vehicle.query.get_or_404(vid)
            db.session.delete(vehicle)
            db.session.commit()
            flash('ลบรถออกจากระบบแล้ว', 'success')

        elif action == 'edit_driver':
            did    = int(request.form.get('driver_id'))
            driver = Driver.query.get_or_404(did)
            driver.name      = request.form.get('name')
            driver.phone     = request.form.get('phone')
            driver.is_active = True if request.form.get('is_active') else False
            driver.user_id   = request.form.get('user_id') or None
            db.session.commit()
            flash(f"อัปเดตข้อมูลคนขับ {driver.name} สำเร็จ!", 'success')

        elif action == 'delete_driver':
            did    = int(request.form.get('driver_id'))
            driver = Driver.query.get_or_404(did)
            db.session.delete(driver)
            db.session.commit()
            flash('ลบพนักงานขับรถออกจากระบบแล้ว', 'success')

        return redirect(url_for('adminfleet.manage_fleet'))

    vehicles = Vehicle.query.order_by(Vehicle.id).all()
    drivers  = Driver.query.order_by(Driver.id).all()
    users    = User.query.order_by(User.full_name).all()
    return render_template('vehicle/admin/admin_manage_fleet.html',
                           vehicles=vehicles, drivers=drivers, users=users,
                           now=datetime.now())


# ─────────────────────────────────────────────
# Admin: หน้าจัดการทริป (รวมทริป / เปลี่ยนรถ)
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin')
@login_required
def admin_trips():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))
 
    bookings = VehicleBooking.query.order_by(VehicleBooking.created_at.desc()).all()
    vehicles = Vehicle.query.filter_by(status='active').order_by(Vehicle.id).all()
    drivers  = Driver.query.filter_by(is_active=True).order_by(Driver.id).all()
 
    # จัดกลุ่ม trip_group เพื่อโชว์ summary
    groups = {}
    for b in bookings:
        if b.trip_group:
            if b.trip_group not in groups:
                groups[b.trip_group] = []
            groups[b.trip_group].append(b)
 
    users_dept = [r[0] for r in db.session.query(DepartmentBudget.department)\
                  .distinct().order_by(DepartmentBudget.department).all()]
 
    return render_template('vehicle/admin/vehicle_admin.html',
                           bookings=bookings,
                           vehicles=vehicles,
                           drivers=drivers,
                           groups=groups,
                           users_dept=users_dept)


# ─────────────────────────────────────────────
# Admin: รวมทริป (Merge)
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/merge', methods=['POST'])
@login_required
def admin_merge():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    booking_ids         = request.form.getlist('booking_ids')
    assigned_vehicle_id = request.form.get('assigned_vehicle_id')
    driver_id           = request.form.get('driver_id') or None
    trip_group          = request.form.get('trip_group', '').strip()
    merge_action        = request.form.get('merge_action', 'approve')  # 'approve' หรือ 'forward'

    if len(booking_ids) < 2:
        flash('กรุณาเลือกรายการอย่างน้อย 2 รายการเพื่อรวมทริป', 'warning')
        return redirect(url_for('vehicle.admin_trips'))

    if not assigned_vehicle_id:
        flash('กรุณาเลือกรถที่จะใช้สำหรับทริปนี้', 'warning')
        return redirect(url_for('vehicle.admin_trips'))

    # เช็คว่ามีรายการที่ขอคนขับแต่ไม่ได้เลือกคนขับ
    need_driver_count = sum(
        1 for bid in booking_ids
        for b in [VehicleBooking.query.get(int(bid))]
        if b and b.need_driver
    )
    if need_driver_count > 0 and not driver_id:
        flash(f'มี {need_driver_count} รายการที่ขอคนขับ กรุณาเลือกคนขับด้วย', 'warning')
        return redirect(url_for('vehicle.admin_trips'))

    # สร้างชื่อกลุ่มอัตโนมัติถ้าไม่ได้กรอก
    if not trip_group:
        count  = db.session.query(VehicleBooking.trip_group)\
                           .filter(VehicleBooking.trip_group.isnot(None))\
                           .distinct().count()
        trip_group = f"TRP-{str(count + 1).zfill(3)}"

    # กำหนด status ตาม action
    new_status = 'approved' if merge_action == 'approve' else 'waiting_approver'

    # อัปเดตทุก booking ที่เลือก
    for bid in booking_ids:
        booking = VehicleBooking.query.get(int(bid))
        if booking:
            booking.trip_group          = trip_group
            booking.assigned_vehicle_id = int(assigned_vehicle_id)
            booking.status              = new_status
            if driver_id and booking.need_driver:
                booking.driver_id = int(driver_id)

    db.session.commit()

    if merge_action == 'forward':
        flash(f'รวมทริป {len(booking_ids)} รายการเป็นกลุ่ม {trip_group} และส่งต่อให้ Approver พิจารณาแล้ว', 'info')
        for bid in booking_ids:
            b = VehicleBooking.query.get(int(bid))
            if b: notify_forwarded_to_approver(b)
    else:
        flash(f'รวมทริป {len(booking_ids)} รายการเป็นกลุ่ม {trip_group} และอนุมัติเรียบร้อยแล้ว', 'success')
        for bid in booking_ids:
            b = VehicleBooking.query.get(int(bid))
            if b: notify_approved(b)
    return redirect(url_for('vehicle.admin_trips'))


# ─────────────────────────────────────────────
# Admin: เปลี่ยนรถ / แก้กลุ่มรายการเดี่ยว
# ─────────────────────────────────────────────
@vehicle_bp.route('/vehicle/admin/assign/<int:booking_id>', methods=['POST'])
@login_required
def admin_assign(booking_id):
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    booking              = VehicleBooking.query.get_or_404(booking_id)
    assigned_vehicle_id  = request.form.get('assigned_vehicle_id')
    assigned_vehicle2_id = request.form.get('assigned_vehicle2_id') or None
    driver_id            = request.form.get('driver_id') or None
    driver2_id           = request.form.get('driver2_id') or None
    trip_group           = request.form.get('trip_group', '').strip() or None
    action               = request.form.get('action', 'assign')
    assign_action        = request.form.get('assign_action', 'approve')

    if action == 'ungroup':
        booking.trip_group           = None
        booking.assigned_vehicle_id  = None
        booking.assigned_vehicle2_id = None
        db.session.commit()
        flash(f'นำ #{booking_id} ออกจากกลุ่มทริปแล้ว', 'success')
    else:
        if booking.need_driver and not driver_id and not booking.driver_id:
            flash(f'รายการ #{booking_id} ขอคนขับ กรุณาเลือกคนขับด้วย', 'warning')
            return redirect(url_for('vehicle.admin_trips'))

        if assigned_vehicle_id:
            booking.assigned_vehicle_id  = int(assigned_vehicle_id)
        booking.assigned_vehicle2_id = int(assigned_vehicle2_id) if assigned_vehicle2_id else None
        if driver_id:
            booking.driver_id  = int(driver_id)
        booking.driver2_id       = int(driver2_id) if driver2_id else None
        booking.trip_group       = trip_group
        booking.expense_type     = request.form.get('expense_type') or None
        booking.central_category = request.form.get('central_category') or None
        booking.trip_department  = request.form.get('trip_department') or booking.user.department or None

        if assign_action == 'forward':
            booking.status = 'waiting_approver'
            db.session.commit()
            notify_forwarded_to_approver(booking)
            flash(f'ส่งต่อรายการ #{booking_id} ให้ Approver แผนก {booking.user.department} พิจารณาแล้ว', 'info')
        else:
            booking.status = 'approved'
            db.session.commit()
            notify_approved(booking)
            flash(f'อัปเดตและอนุมัติรายการ #{booking_id} เรียบร้อยแล้ว', 'success')

    return redirect(url_for('vehicle.admin_trips'))

# ─────────────────────────────────────────────
# กรอกไมล์ (คนขับ + superadmin)
# ─────────────────────────────────────────────
def is_driver():
    return current_user.role_vehicle == 'driver' or current_user.is_superadmin

@vehicle_bp.route('/vehicle/mileage', methods=['GET', 'POST'])
@login_required
def mileage_log():
    if not is_driver():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    if request.method == 'POST':
        booking_id = int(request.form.get('booking_id'))
        booking    = VehicleBooking.query.get_or_404(booking_id)
        entry_type = request.form.get('entry_type')

        mileage = VehicleMileage.query.filter_by(booking_id=booking_id).first()
        if not mileage:
            mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
            db.session.add(mileage)

        upload_folder = os.path.join('static', 'uploads', 'mileage')
        os.makedirs(upload_folder, exist_ok=True)

        if entry_type == 'start':
            mileage.odometer_start = int(request.form.get('odometer_start', 0))
            mileage.actual_start   = datetime.strptime(request.form.get('actual_start'), '%Y-%m-%dT%H:%M')
            # รูปหน้าปัดก่อนออก
            img = request.files.get('odometer_start_img')
            if img and img.filename:
                fname = f"{int(time.time())}_start_{secure_filename(img.filename)}"
                img.save(os.path.join(upload_folder, fname))
                mileage.odometer_start_img = fname
            flash(f'บันทึกเลขไมล์ก่อนออก #{booking_id} เรียบร้อย', 'success')

        elif entry_type == 'end':
            submitted_end_mileage = int(request.form.get('odometer_end', 0))
            # 🌟 เช็คว่าเลขไมล์ตอนจบ ต้องมากกว่าเลขไมล์ตอนเริ่ม
            # (ดักเผื่อกรณี mileage.odometer_start มีค่าอยู่แล้ว)
            if mileage.odometer_start is not None and submitted_end_mileage <= mileage.odometer_start:
                flash(f'❌ บันทึกไม่สำเร็จ! เลขไมล์ตอนจบ ({submitted_end_mileage}) ต้องมากกว่าเลขไมล์ตอนเริ่ม ({mileage.odometer_start})', 'danger')
                return redirect(url_for('vehicle.mileage_log')) # เด้งกลับไปให้กรอกใหม่
            
            # ถ้าเลขไมล์ถูกต้อง ค่อยเอาไปใส่ใน object
            mileage.odometer_end = submitted_end_mileage
            mileage.actual_end   = datetime.strptime(request.form.get('actual_end'), '%Y-%m-%dT%H:%M')
            # รูปหน้าปัดหลังกลับ
            img = request.files.get('odometer_end_img')
            if img and img.filename:
                fname = f"{int(time.time())}_end_{secure_filename(img.filename)}"
                img.save(os.path.join(upload_folder, fname))
                mileage.odometer_end_img = fname
            # เติมน้ำมันระหว่างทาง
            mileage.refuel = True if request.form.get('refuel') else False
            if mileage.refuel:
                refuel_amt = request.form.get('refuel_amount', '').strip()
                if refuel_amt:
                    mileage.refuel_amount = float(refuel_amt)
                refuel_img = request.files.get('refuel_img')
                if refuel_img and refuel_img.filename:
                    fname = f"{int(time.time())}_refuel_{secure_filename(refuel_img.filename)}"
                    refuel_img.save(os.path.join(upload_folder, fname))
                    mileage.refuel_img = fname
            # admin กรอกค่าน้ำมัน manual
            fuel = request.form.get('fuel_cost', '').strip()
            if fuel:
                mileage.fuel_cost = float(fuel)
            flash(f'บันทึกเลขไมล์หลังกลับ #{booking_id} เรียบร้อย', 'success')

        db.session.commit()

        # หักงบแผนกอัตโนมัติ
        if entry_type == 'end' and booking.trip_department and booking.expense_type == 'department':
            m2         = VehicleMileage.query.filter_by(booking_id=booking_id).first()
            distance   = (m2.odometer_end - m2.odometer_start) if (m2 and m2.odometer_end and m2.odometer_start) else None
            fuel_price = float(SystemConfig.get('fuel_price', '40'))
            trip_cost  = m2.fuel_cost if (m2 and m2.fuel_cost) else \
                         (round((distance / booking.assigned_vehicle.fuel_rate) * fuel_price, 2)
                          if distance and booking.assigned_vehicle and booking.assigned_vehicle.fuel_rate else 0)
            if trip_cost > 0:
                now2   = datetime.now()
                budget = DepartmentBudget.query.filter_by(
                    department=booking.trip_department,
                    year=now2.year, month=now2.month
                ).first()
                if budget:
                    budget.used_amount = round(budget.used_amount + trip_cost, 2)
                    db.session.commit()

        return redirect(url_for('vehicle.mileage_log'))

    bookings = VehicleBooking.query.filter_by(status='approved')\
                                   .order_by(VehicleBooking.start_datetime.desc()).all()
    return render_template('vehicle/vehicle_mileage.html', bookings=bookings)


# ─────────────────────────────────────────────
# สรุปค่าใช้จ่าย (admin + superadmin)
# ─────────────────────────────────────────────
WORK_START = 8    # 08:00
WORK_END   = 16   # 16:00
OT_RATE    = 20   # บาท/ชม
SUN_FLAT   = 300  # บาท (วันอาทิตย์)

def calc_ot(booking, mileage):
    """คำนวณค่า OT คนขับ"""
    if not booking.need_driver or not mileage or not mileage.actual_end:
        return 0

    actual_end = mileage.actual_end
    weekday    = actual_end.weekday()  # 0=จันทร์ … 6=อาทิตย์

    if weekday == 6:  # อาทิตย์
        return SUN_FLAT

    # จันทร์-เสาร์: นับ OT หลัง 16:00 เท่านั้น
    if actual_end.hour < WORK_END:
        return 0

    ot_minutes = (actual_end.hour - WORK_END) * 60 + actual_end.minute
    ot_hours   = ot_minutes / 60
    return round(ot_hours * OT_RATE, 2)


@admincost_bp.route('/vehicle/mileage/override-fuel', methods=['POST'])
@login_required
def override_fuel():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
    booking_id = int(request.form.get('booking_id'))
    fuel_cost  = float(request.form.get('fuel_cost', 0))
    mileage = VehicleMileage.query.filter_by(booking_id=booking_id).first()
    if not mileage:
        mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
        db.session.add(mileage)
    mileage.fuel_cost = fuel_cost
    db.session.commit()
    flash(f'Override ค่าน้ำมัน #{booking_id} เป็น {fuel_cost:,.2f} บาท เรียบร้อย', 'success')
    return redirect(request.referrer or url_for('admincost.cost_summary'))


@admincost_bp.route('/admin/cost', methods=['GET', 'POST'])
@login_required
def cost_summary():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์เข้าหน้านี้', 'danger')
        return redirect(url_for('vehicle.index'))

    # บันทึกราคาน้ำมัน
    if request.method == 'POST' and request.form.get('action') == 'save_fuel_price':
        new_price = request.form.get('fuel_price', '').strip()
        if new_price:
            SystemConfig.set('fuel_price', new_price)
            flash(f'อัปเดตราคาน้ำมันเป็น {new_price} บาท/ลิตร แล้ว', 'success')
        return redirect(url_for('admincost.cost_summary'))

    # ราคาน้ำมันปัจจุบัน
    fuel_price = float(SystemConfig.get('fuel_price', '40'))

    # filter เดือน/ปี
    now         = datetime.now()
    sel_year    = int(request.args.get('year',  now.year))
    sel_month   = int(request.args.get('month', now.month))
    sel_expense = request.args.get('expense_type', '')

    q = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        extract('year',  VehicleBooking.start_datetime) == sel_year,
        extract('month', VehicleBooking.start_datetime) == sel_month,
    )
    if sel_expense:
        q = q.filter(VehicleBooking.expense_type == sel_expense)

    bookings = q.order_by(VehicleBooking.start_datetime).all()

    rows = []
    totals = {'central': 0, 'department': 0, 'personal': 0}

    for b in bookings:
        m        = b.mileage[0] if b.mileage else None
        distance = (m.odometer_end - m.odometer_start) if (m and m.odometer_end and m.odometer_start) else None

        # คำนวณค่าน้ำมัน — ถ้ามี override ใช้เลย ไม่งั้นคำนวณจาก odometer
        if m and m.fuel_cost:
            fuel = m.fuel_cost
        elif distance and b.assigned_vehicle and b.assigned_vehicle.fuel_rate:
            fuel = round((distance / b.assigned_vehicle.fuel_rate) * fuel_price, 2)
        else:
            fuel = 0

        ot       = calc_ot(b, m)
        total    = (fuel or 0) + ot
        exp      = b.expense_type or 'unknown'
        if exp in totals:
            totals[exp] += total
        rows.append({
            'booking':  b,
            'mileage':  m,
            'distance': distance,
            'fuel':     fuel,
            'ot':       ot,
            'total':    total,
        })

    TH_MONTHS = ['', 'ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.',
                 'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
    month_label = f"{TH_MONTHS[sel_month]} {sel_year + 543}"

    return render_template('vehicle/admin/vehicle_cost.html',
                           rows=rows, totals=totals,
                           sel_year=sel_year, sel_month=sel_month,
                           sel_expense=sel_expense,
                           month_label=month_label,
                           fuel_price=fuel_price,
                           now=now)

# ─────────────────────────────────────────────
# Driver View — หน้าคนขับ (mobile-friendly)
# ─────────────────────────────────────────────
@driver_bp.route('/driver')
@login_required
def driver_home():
    # หา Driver record ที่ผูกกับ user นี้
    driver = Driver.query.filter_by(user_id=current_user.id).first()
    if not driver:
        flash('บัญชีของคุณยังไม่ได้ผูกกับพนักงานขับรถ กรุณาติดต่อ Admin', 'warning')
        return redirect(url_for('vehicle.index'))

    # ดึงทริปที่ approved และคนขับคือตัวเอง (slot 1 หรือ slot 2)
    from sqlalchemy import or_
    bookings = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        or_(
            VehicleBooking.driver_id  == driver.id,
            VehicleBooking.driver2_id == driver.id
        )
    ).order_by(VehicleBooking.start_datetime.desc()).all()

    return render_template('vehicle/driver_home.html',
                           driver=driver,
                           bookings=bookings,
                           today_start=datetime.now().replace(hour=0, minute=0, second=0),
                           today_end=datetime.now().replace(hour=23, minute=59, second=59))


@driver_bp.route('/driver/mileage', methods=['POST'])
@login_required
def driver_mileage():
    driver = Driver.query.filter_by(user_id=current_user.id).first()
    if not driver:
        flash('ไม่พบข้อมูลคนขับ', 'danger')
        return redirect(url_for('driver.driver_home'))

    booking_id = int(request.form.get('booking_id'))
    booking    = VehicleBooking.query.get_or_404(booking_id)

    # ตรวจสอบว่าทริปนี้เป็นของคนขับคนนี้จริง
    if booking.driver_id != driver.id and booking.driver2_id != driver.id:
        flash('คุณไม่มีสิทธิ์บันทึกทริปนี้', 'danger')
        return redirect(url_for('driver.driver_home'))

    entry_type = request.form.get('entry_type')
    mileage    = VehicleMileage.query.filter_by(booking_id=booking_id).first()
    if not mileage:
        mileage = VehicleMileage(booking_id=booking_id, noted_by=current_user.id)
        db.session.add(mileage)

    upload_folder = os.path.join('static', 'uploads', 'mileage')
    os.makedirs(upload_folder, exist_ok=True)

    if entry_type == 'start':
        mileage.odometer_start = int(request.form.get('odometer_start', 0))
        mileage.actual_start   = datetime.strptime(request.form.get('actual_start'), '%Y-%m-%dT%H:%M')
        img = request.files.get('odometer_start_img')
        if img and img.filename:
            fname = f"{int(time.time())}_start_{secure_filename(img.filename)}"
            img.save(os.path.join(upload_folder, fname))
            mileage.odometer_start_img = fname
        flash('✅ บันทึกเลขไมล์ก่อนออกเรียบร้อย', 'success')

    elif entry_type == 'end':
        submitted_end_mileage = int(request.form.get('odometer_end', 0))
        # 🌟 เช็คว่าเลขไมล์ตอนจบ ต้องมากกว่าเลขไมล์ตอนเริ่ม
        # (ดักเผื่อกรณี mileage.odometer_start มีค่าอยู่แล้ว)
        if mileage.odometer_start is not None and submitted_end_mileage <= mileage.odometer_start:
            flash(f'❌ บันทึกไม่สำเร็จ! เลขไมล์ตอนจบ ({submitted_end_mileage}) ต้องมากกว่าเลขไมล์ตอนเริ่ม ({mileage.odometer_start})', 'danger')
            return redirect(url_for('driver.driver_home')) # เด้งกลับไปให้กรอกใหม่
        
        # ถ้าเลขไมล์ถูกต้อง ค่อยเอาไปใส่ใน object
        mileage.odometer_end = submitted_end_mileage
        mileage.actual_end   = datetime.strptime(request.form.get('actual_end'), '%Y-%m-%dT%H:%M')
        img = request.files.get('odometer_end_img')
        if img and img.filename:
            fname = f"{int(time.time())}_end_{secure_filename(img.filename)}"
            img.save(os.path.join(upload_folder, fname))
            mileage.odometer_end_img = fname
        mileage.refuel = True if request.form.get('refuel') else False
        if mileage.refuel:
            amt = request.form.get('refuel_amount', '').strip()
            if amt:
                mileage.refuel_amount = float(amt)
            ri = request.files.get('refuel_img')
            if ri and ri.filename:
                fname = f"{int(time.time())}_refuel_{secure_filename(ri.filename)}"
                ri.save(os.path.join(upload_folder, fname))
                mileage.refuel_img = fname
        flash('✅ ปิดงานเรียบร้อย', 'success')

    db.session.commit()

    # ── หักงบประมาณแผนกอัตโนมัติเมื่อปิดงาน ──
    if entry_type == 'end' and booking.trip_department and booking.expense_type == 'department':
        m2       = VehicleMileage.query.filter_by(booking_id=booking_id).first()
        distance = (m2.odometer_end - m2.odometer_start) if (m2 and m2.odometer_end and m2.odometer_start) else None
        fuel_price = float(SystemConfig.get('fuel_price', '40'))
        if m2 and m2.fuel_cost:
            trip_cost = m2.fuel_cost
        elif distance and booking.assigned_vehicle and booking.assigned_vehicle.fuel_rate:
            trip_cost = round((distance / booking.assigned_vehicle.fuel_rate) * fuel_price, 2)
        else:
            trip_cost = 0
        if trip_cost > 0:
            now2   = datetime.now()
            budget = DepartmentBudget.query.filter_by(
                department=booking.trip_department,
                year=now2.year, month=now2.month
            ).first()
            if budget:
                budget.used_amount = round(budget.used_amount + trip_cost, 2)
                db.session.commit()

    return redirect(url_for('driver.driver_home'))


# ══════════════════════════════════════════════════════
# Feature 3: Budget Routes
# ══════════════════════════════════════════════════════
@adminfleet_bp.route('/admin/budget', methods=['GET', 'POST'])
@login_required
def budget_manage():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))
 
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'set_budget':
            dept   = request.form.get('department', '').strip()
            year   = int(request.form.get('year'))
            month  = int(request.form.get('month'))
            amount = float(request.form.get('budget_amount', 0))
            budget = DepartmentBudget.query.filter_by(department=dept, year=year, month=month).first()
            if budget:
                budget.budget_amount = amount
            else:
                budget = DepartmentBudget(department=dept, year=year, month=month, budget_amount=amount)
                db.session.add(budget)
            db.session.commit()
            flash(f'ตั้งงบ {dept} เดือน {month}/{year} = {amount:,.0f} บาท เรียบร้อย', 'success')
        return redirect(url_for('adminfleet.budget_manage'))
 
    now        = datetime.now()
    sel_year   = int(request.args.get('year', now.year))
    sel_month  = int(request.args.get('month', now.month))
    raw_budgets = DepartmentBudget.query.filter_by(year=sel_year, month=sel_month)\
                                        .order_by(DepartmentBudget.department).all()
 
    # คำนวณ pct และ remaining ใน Python ก่อนส่งไป template
    budgets = []
    for b in raw_budgets:
        pct = round(min(b.used_amount / b.budget_amount * 100, 100), 1) if b.budget_amount > 0 else 0
        budgets.append({
            'id':            b.id,
            'department':    b.department,
            'budget_amount': b.budget_amount,
            'used_amount':   b.used_amount,
            'remaining':     round(b.budget_amount - b.used_amount, 2),
            'pct':           pct,
        })
 
    departments = [r[0] for r in db.session.query(User.department).filter(
        User.department.isnot(None)).distinct().order_by(User.department).all()]
 
    TH_MONTHS = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
    return render_template('vehicle/admin/budget_manage.html',
                           budgets=budgets, departments=departments,
                           sel_year=sel_year, sel_month=sel_month,
                           month_label=f"{TH_MONTHS[sel_month]} {sel_year+543}",
                           now=now)


# ══════════════════════════════════════════════════════
# Feature 4: Vehicle History (API — ใช้ใน manage-fleet)
# ══════════════════════════════════════════════════════
@adminfleet_bp.route('/api/vehicle/<int:vid>/history')
@login_required
def vehicle_history(vid):
    vehicle  = Vehicle.query.get_or_404(vid)
    bookings = VehicleBooking.query.filter(
        VehicleBooking.assigned_vehicle_id == vid,
        VehicleBooking.status == 'approved'
    ).order_by(VehicleBooking.start_datetime.desc()).limit(20).all()

    rows = []
    total_km = 0
    for b in bookings:
        m = b.mileage[0] if b.mileage else None
        dist = (m.odometer_end - m.odometer_start) if (m and m.odometer_end and m.odometer_start) else None
        if dist:
            total_km += dist
        rows.append({
            'id':          b.id,
            'date':        b.start_datetime.strftime('%d/%m/%Y'),
            'destination': b.destination,
            'driver':      b.driver.name if b.driver else '-',
            'distance':    dist,
            'odometer_end': m.odometer_end if m else None,
        })
    return jsonify({'vehicle': f"{vehicle.brand} {vehicle.model}", 'total_km': total_km, 'rows': rows})


# ══════════════════════════════════════════════════════
# Feature 5: Excel Export
# ══════════════════════════════════════════════════════
@admincost_bp.route('/admin/cost/export')
@login_required
def cost_export():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('vehicle.index'))

    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('ไม่พบ openpyxl — รัน: pip install openpyxl', 'danger')
        return redirect(url_for('admincost.cost_summary'))

    from flask import send_file
    now       = datetime.now()
    sel_year  = int(request.args.get('year', now.year))
    sel_month = int(request.args.get('month', now.month))
    sel_exp   = request.args.get('expense_type', '')
    fuel_price = float(SystemConfig.get('fuel_price', '40'))

    q = VehicleBooking.query.filter(
        VehicleBooking.status == 'approved',
        extract('year',  VehicleBooking.start_datetime) == sel_year,
        extract('month', VehicleBooking.start_datetime) == sel_month,
    )
    if sel_exp:
        q = q.filter(VehicleBooking.expense_type == sel_exp)
    bookings = q.order_by(VehicleBooking.start_datetime).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    TH_MONTHS = ['','ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
    ws.title = f"ค่าใช้จ่าย {TH_MONTHS[sel_month]} {sel_year+543}"

    # Header row
    headers = ['#','ผู้จอง','แผนก','ปลายทาง','รถ','คนขับ','วันเดินทาง',
               'ไมล์เริ่ม','ไมล์จบ','ระยะทาง(กม.)','ค่าน้ำมัน(฿)','OT(฿)','รวม(฿)','ประเภท']
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF', name='Sarabun')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font       = hdr_font
        cell.fill       = hdr_fill
        cell.alignment  = Alignment(horizontal='center', vertical='center')
        cell.border     = border

    EXP_LABEL = {'central':'ส่วนกลาง','department':'หน่วยงาน','personal':'ส่วนตัว'}
    total_fuel = total_ot = total_all = 0

    for ri, b in enumerate(bookings, 2):
        m        = b.mileage[0] if b.mileage else None
        distance = (m.odometer_end - m.odometer_start) if (m and m.odometer_end and m.odometer_start) else None
        if m and m.fuel_cost:
            fuel = m.fuel_cost
        elif distance and b.assigned_vehicle and b.assigned_vehicle.fuel_rate:
            fuel = round((distance / b.assigned_vehicle.fuel_rate) * fuel_price, 2)
        else:
            fuel = 0
        ot    = calc_ot(b, m)
        total = fuel + ot
        total_fuel += fuel; total_ot += ot; total_all += total

        row_data = [
            b.id,
            b.user.full_name or b.user.username,
            b.trip_department or b.user.department or '-',
            b.destination,
            f"{b.assigned_vehicle.brand} {b.assigned_vehicle.model} ({b.assigned_vehicle.license_plate})" if b.assigned_vehicle else '-',
            b.driver.name if b.driver else '-',
            b.start_datetime.strftime('%d/%m/%Y'),
            m.odometer_start if m else None,
            m.odometer_end if m else None,
            distance,
            round(fuel, 2),
            round(ot, 2),
            round(total, 2),
            EXP_LABEL.get(b.expense_type or '', 'ไม่ระบุ'),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal='center' if ci in [1,7,8,9,10,11,12,13] else 'left')
            if ri % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')

    # Total row
    tr = len(bookings) + 2
    ws.cell(row=tr, column=10, value='รวม').font = Font(bold=True)
    ws.cell(row=tr, column=11, value=round(total_fuel, 2)).font = Font(bold=True)
    ws.cell(row=tr, column=12, value=round(total_ot, 2)).font = Font(bold=True)
    ws.cell(row=tr, column=13, value=round(total_all, 2)).font = Font(bold=True)

    col_widths = [6,20,18,22,24,16,14,12,12,14,14,10,12,12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"vehicle_cost_{sel_year}_{sel_month:02d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════
# Feature 6: Vehicle Service/Tax Date Update
# ══════════════════════════════════════════════════════
@adminfleet_bp.route('/admin/manage-fleet/service', methods=['POST'])
@login_required
def update_vehicle_service():
    if not is_vehicle_admin():
        flash('คุณไม่มีสิทธิ์', 'danger')
        return redirect(url_for('adminfleet.manage_fleet'))

    vid     = int(request.form.get('vehicle_id'))
    vehicle = Vehicle.query.get_or_404(vid)

    svc_date = request.form.get('next_service_date', '').strip()
    svc_km   = request.form.get('next_service_km', '').strip()
    tax_date = request.form.get('tax_due_date', '').strip()

    vehicle.next_service_date = date.fromisoformat(svc_date) if svc_date else None
    vehicle.next_service_km   = int(svc_km) if svc_km else None
    vehicle.tax_due_date      = date.fromisoformat(tax_date) if tax_date else None
    db.session.commit()
    flash(f'อัปเดตวันนัดซ่อม/ต่อภาษี {vehicle.brand} {vehicle.model} เรียบร้อย', 'success')
    return redirect(url_for('adminfleet.manage_fleet'))