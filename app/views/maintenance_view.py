import os
import io
import time
from datetime import datetime
from werkzeug.utils import secure_filename
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from sqlalchemy import extract, func
from models import db, MaintenanceTicket

maintenance_bp = Blueprint('maintenance', __name__)


def is_maintenance_admin():
    """เช็คว่า user คนนี้มีสิทธิ์ admin ระบบซ่อมอาคารหรือเปล่า"""
    return current_user.role_maintenance == 'admin' or current_user.is_superadmin


def get_summary_context():
    """Helper: คำนวณ summary เดือนปัจจุบัน + available_months สำหรับ Export"""
    now = datetime.now()
    current_month_tickets = MaintenanceTicket.query.filter(
        extract('year',  MaintenanceTicket.created_at) == now.year,
        extract('month', MaintenanceTicket.created_at) == now.month
    ).all()

    summary = {
        'month_label': now.strftime('%B %Y'),
        'total':       len(current_month_tickets),
        'pending':     sum(1 for t in current_month_tickets if t.status == 'pending'),
        'in_progress': sum(1 for t in current_month_tickets if t.status == 'in_progress'),
        'done':        sum(1 for t in current_month_tickets if t.status == 'done'),
        'total_cost':  sum(float(t.repair_cost or 0) for t in current_month_tickets if t.status == 'done'),
        'by_category': {},
    }
    for t in current_month_tickets:
        summary['by_category'][t.category] = summary['by_category'].get(t.category, 0) + 1

    raw_months = db.session.query(
        func.strftime('%Y-%m', MaintenanceTicket.created_at).label('month')
    ).distinct().order_by('month').all()
    available_months = [m.month for m in raw_months if m.month]

    return summary, available_months


def save_image(file, subfolder):
    """Helper: บันทึกไฟล์รูปภาพ คืนค่าชื่อไฟล์"""
    if file and file.filename != '':
        upload_folder = os.path.join('static', 'uploads', subfolder)
        os.makedirs(upload_folder, exist_ok=True)
        safe_name = secure_filename(file.filename)
        filename = f"{int(time.time())}_{safe_name}"
        file.save(os.path.join(upload_folder, filename))
        return filename
    return None


@maintenance_bp.route('/maintenance', methods=['GET', 'POST'])
@login_required
def index():
    if request.method == 'POST':
        filename = save_image(request.files.get('image'), 'maintenance')

        new_ticket = MaintenanceTicket(
            user_id=current_user.id,
            category=request.form.get('category'),
            urgency=request.form.get('urgency'),
            location=request.form.get('location'),
            contact_number=request.form.get('contact_number'),
            subject=request.form.get('subject'),
            image_file=filename
        )
        db.session.add(new_ticket)
        db.session.commit()

        flash('แจ้งซ่อมสำเร็จ! ส่งเรื่องให้ช่างอาคารสถานที่เรียบร้อยแล้ว', 'success')
        return redirect(url_for('maintenance.index'))

    tickets = MaintenanceTicket.query.order_by(MaintenanceTicket.created_at.desc()).all()
    summary, available_months = get_summary_context()

    return render_template(
        'maintenance/maintenance.html',
        tickets=tickets,
        summary=summary,
        available_months=available_months
    )


# 🟢 Route แก้ไข — เฉพาะเจ้าของ หรือ admin, เฉพาะ status pending
@maintenance_bp.route('/maintenance/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit(id):
    ticket = MaintenanceTicket.query.get_or_404(id)

    if ticket.user_id != current_user.id and not is_maintenance_admin():
        flash('คุณไม่มีสิทธิ์แก้ไขรายการนี้', 'danger')
        return redirect(url_for('maintenance.index'))

    if request.method == 'POST':
        ticket.category       = request.form.get('category')
        ticket.urgency        = request.form.get('urgency')
        ticket.location       = request.form.get('location')
        ticket.contact_number = request.form.get('contact_number')
        ticket.subject        = request.form.get('subject')

        db.session.commit()
        flash('อัปเดตข้อมูลแจ้งซ่อมเรียบร้อยแล้ว', 'success')
        return redirect(url_for('maintenance.index'))

    tickets = MaintenanceTicket.query.order_by(MaintenanceTicket.created_at.desc()).all()
    summary, available_months = get_summary_context()
    return render_template('maintenance/maintenance.html', tickets=tickets, edit_ticket=ticket, summary=summary, available_months=available_months)


# 🟢 Route ลบ — เฉพาะเจ้าของ หรือ admin
@maintenance_bp.route('/maintenance/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    ticket = MaintenanceTicket.query.get_or_404(id)

    if ticket.user_id != current_user.id and not is_maintenance_admin():
        flash('คุณไม่มีสิทธิ์ลบรายการนี้', 'danger')
        return redirect(url_for('maintenance.index'))

    db.session.delete(ticket)
    db.session.commit()
    flash('ลบรายการแจ้งซ่อมเรียบร้อยแล้ว', 'success')
    return redirect(url_for('maintenance.index'))


# 🆕 Route Admin — รับงาน / ปิดงาน
@maintenance_bp.route('/maintenance/update_status/<int:id>', methods=['POST'])
@login_required
def update_status(id):
    if not is_maintenance_admin():
        flash('คุณไม่มีสิทธิ์ดำเนินการนี้', 'danger')
        return redirect(url_for('maintenance.index'))

    ticket = MaintenanceTicket.query.get_or_404(id)
    action = request.form.get('action')

    if action == 'accept':
        if ticket.status != 'pending':
            flash('ไม่สามารถรับงานได้ เนื่องจากสถานะไม่ใช่ "รอดำเนินการ"', 'warning')
            return redirect(url_for('maintenance.index'))

        ticket.status = 'in_progress'

        # กำหนดวันนัดซ่อม (optional)
        scheduled_date_str = request.form.get('scheduled_date', '').strip()
        if scheduled_date_str:
            try:
                ticket.scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        flash(f'รับงาน #{ticket.id} เรียบร้อยแล้ว', 'success')

    elif action == 'close':
        if ticket.status != 'in_progress':
            flash('ไม่สามารถปิดงานได้ เนื่องจากสถานะไม่ใช่ "กำลังดำเนินการ"', 'warning')
            return redirect(url_for('maintenance.index'))

        resolved_note = request.form.get('resolved_note', '').strip()
        if not resolved_note:
            flash('กรุณากรอกบันทึกผลการซ่อมก่อนปิดงาน', 'warning')
            return redirect(url_for('maintenance.index'))

        ticket.status          = 'done'
        ticket.resolved_note   = resolved_note
        ticket.resolved_at     = datetime.now()
        ticket.technician_type = request.form.get('technician_type')

        # ค่าใช้จ่าย
        cost_str = request.form.get('repair_cost', '').strip()
        if cost_str:
            try:
                ticket.repair_cost = float(cost_str)
            except ValueError:
                ticket.repair_cost = None

        # รูปหลังซ่อม (optional)
        after_img = save_image(request.files.get('image_after'), 'maintenance')
        if after_img:
            ticket.image_after = after_img

        flash(f'ปิดงาน #{ticket.id} เรียบร้อยแล้ว', 'success')

    else:
        flash('คำสั่งไม่ถูกต้อง', 'danger')
        return redirect(url_for('maintenance.index'))

    db.session.commit()
    return redirect(url_for('maintenance.index'))


# 🆕 Route Export Excel
@maintenance_bp.route('/maintenance/export_excel')
@login_required
def export_excel():
    if not is_maintenance_admin():
        flash('คุณไม่มีสิทธิ์ Export ข้อมูล', 'danger')
        return redirect(url_for('maintenance.index'))

    month_str = request.args.get('month', '')  # format: YYYY-MM
    if not month_str:
        flash('กรุณาระบุเดือนที่ต้องการ Export', 'warning')
        return redirect(url_for('maintenance.index'))

    try:
        year, month = int(month_str[:4]), int(month_str[5:7])
    except (ValueError, IndexError):
        flash('รูปแบบเดือนไม่ถูกต้อง', 'danger')
        return redirect(url_for('maintenance.index'))

    tickets = MaintenanceTicket.query.filter(
        extract('year',  MaintenanceTicket.created_at) == year,
        extract('month', MaintenanceTicket.created_at) == month
    ).order_by(MaintenanceTicket.created_at.asc()).all()

    # สร้าง Excel ด้วย openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('ไม่พบ library openpyxl กรุณาติดตั้งด้วย: pip install openpyxl', 'danger')
        return redirect(url_for('maintenance.index'))

    wb = openpyxl.Workbook()
    ws = wb.active

    THAI_MONTHS = {
        1:'มกราคม', 2:'กุมภาพันธ์', 3:'มีนาคม',    4:'เมษายน',
        5:'พฤษภาคม', 6:'มิถุนายน',  7:'กรกฎาคม',   8:'สิงหาคม',
        9:'กันยายน', 10:'ตุลาคม',   11:'พฤศจิกายน', 12:'ธันวาคม'
    }
    month_label = f"{THAI_MONTHS.get(month, month)} {year + 543}"
    ws.title = month_label[:31]

    # --- สไตล์ ---
    header_fill   = PatternFill("solid", fgColor="F59E0B")   # amber
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    title_font    = Font(bold=True, size=13)
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # --- Title Row ---
    ws.merge_cells('A1:L1')
    ws['A1'] = f'รายงานการแจ้งซ่อมอาคารสถานที่ — {month_label}'
    ws['A1'].font      = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:L2')
    ws['A2'] = f'จำนวนรายการทั้งหมด: {len(tickets)} รายการ  |  ค่าใช้จ่ายรวม: {sum(float(t.repair_cost or 0) for t in tickets if t.status == "done"):,.2f} บาท'
    ws['A2'].font      = Font(italic=True, size=10, color="555555")
    ws['A2'].alignment = center_align

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # --- Header Row ---
    headers = [
        'ID', 'วันที่แจ้ง', 'ผู้แจ้ง', 'ประเภทงาน', 'สถานที่',
        'เบอร์ติดต่อ', 'ความเร่งด่วน', 'หัวข้อปัญหา', 'สถานะ',
        'ประเภทช่าง', 'ค่าใช้จ่าย (บาท)', 'บันทึกการซ่อม'
    ]
    col_widths = [6, 16, 16, 18, 20, 14, 12, 28, 12, 12, 16, 35]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 22

    # --- Data Rows ---
    STATUS_MAP   = {'pending': 'รอดำเนินการ', 'in_progress': 'กำลังดำเนินการ', 'done': 'เสร็จสิ้น'}
    done_fill    = PatternFill("solid", fgColor="D1FAE5")  # green-100
    pending_fill = PatternFill("solid", fgColor="F3F4F6")  # gray-100

    total_cost = 0.0
    for row_idx, t in enumerate(tickets, start=4):
        cost_val = float(t.repair_cost) if t.repair_cost else None
        if cost_val:
            total_cost += cost_val

        row_data = [
            f'#{t.id}',
            t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            t.user.full_name or t.user.username,
            t.category,
            t.location,
            t.contact_number or '',
            t.urgency,
            t.subject,
            STATUS_MAP.get(t.status, t.status),
            t.technician_type or '',
            cost_val,
            t.resolved_note or '',
        ]

        row_fill = done_fill if t.status == 'done' else pending_fill

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.alignment = center_align if col_idx in [1, 2, 6, 7, 8, 9, 10, 11] else left_align
            cell.fill      = row_fill

            # Format ตัวเลขค่าใช้จ่าย
            if col_idx == 11 and value is not None:
                cell.number_format = '#,##0.00'

        ws.row_dimensions[row_idx].height = 18

    # --- Summary Row ---
    sum_row = len(tickets) + 4
    ws.merge_cells(f'A{sum_row}:J{sum_row}')
    ws[f'A{sum_row}'] = 'รวมค่าใช้จ่ายทั้งหมด'
    ws[f'A{sum_row}'].font      = Font(bold=True)
    ws[f'A{sum_row}'].alignment = Alignment(horizontal='right')
    ws[f'A{sum_row}'].fill      = PatternFill("solid", fgColor="FEF3C7")

    ws[f'K{sum_row}'] = total_cost
    ws[f'K{sum_row}'].number_format = '#,##0.00'
    ws[f'K{sum_row}'].font       = Font(bold=True)
    ws[f'K{sum_row}'].alignment  = center_align
    ws[f'K{sum_row}'].fill       = PatternFill("solid", fgColor="FEF3C7")
    ws[f'K{sum_row}'].border     = thin_border

    ws.freeze_panes = 'A4'

    # --- ส่งไฟล์กลับ ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename_export = f"maintenance_{year}_{month:02d}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename_export,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )